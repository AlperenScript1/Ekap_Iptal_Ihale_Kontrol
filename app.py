from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
import chromedriver_autoinstaller
from selenium import webdriver 
from time import sleep, time
from openpyxl import load_workbook 
from openpyxl.styles import PatternFill


driver = webdriver.Chrome(); 
url = "https://ekapv2.kik.gov.tr/ekap/search";
#option = option() 
#option.add_argument("--headless")
driver.set_window_size(1920,770)
driver.get(url);

path = "excel/iptal.xlsx"
wb = load_workbook(path)
ws = wb.active

yesil = PatternFill("solid", fgColor="00FF00")
kirmizi = PatternFill("solid", fgColor="FF0000")
gri = PatternFill("solid", fgColor="D9D9D9")

try:
    iptal = [] #iptal ihalelerin tutulacağı array
    print("================================================================================================")
    for row in ws.iter_rows(min_row=1, max_col=1):
        sleep(1)
        boya = row[0]
        deger = boya.value

        textBox = sonucilanText = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "dx-text-box[id='search-by-word'] input[placeholder='Ara']")))
        textBox.clear()
        if deger is None: #! aratılacak id kalmaz ise işlemler durdurulur ve iptal olarak alınan id'ler yazılır
            print("ihaleler bitti")
            print("İptal olanlar: " + str(iptal))
            driver.quit()
            break

        textBox.send_keys(deger)
        print("find element: textBox")

        searchButton = sonucilanText = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CLASS_NAME, "search-button")))        
        searchButton.click()
        print("find element: searchButton")
        
        #ihale var mı ? kontrolü 
        try:
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//div[@class='grid-result-title']"))).text
            print("ihale sistemde yok")
            boya.fill = kirmizi
        except:
            print("ihale bulundu")
            boya.fill = yesil
      

except ValueError as e :
    print("Hata kodu: " + str(e))


sleep(5)
wb.save("iptalTest.xlsx")
sleep(2)
driver.quit(); 